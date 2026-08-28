"""
Views de agendamentos de salas e dispositivos móveis, concorrência e atribuição de alunos.
"""

import calendar
import uuid
from collections import defaultdict
from datetime import date, timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import IntegrityError, transaction
from django.db.models import Count, Max, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.models import (
    Agendamento,
    Equipamento,
    ItemDispositivo,
    Notificacao,
    RelacaoAlunoEquipamento,
    Sala,
    Turma,
)
from app.services.historico_service import registrar_acao
from .common import (
    AULAS_HORARIOS,
    DIAS_SEMANA_LONGO,
    DIAS_SEMANA_PT,
    MESES_PT,
    is_admin_aprovado,
    is_ajax,
    is_usuario_aprovado,
)


def _resolver_professor(request):
    is_admin = is_admin_aprovado(request.user)
    prof_id = request.POST.get('professor')
    if is_admin and prof_id:
        prof = User.objects.filter(id=prof_id).first()
        if prof:
            return prof
    return request.user


def _professores_aprovados():
    return User.objects.filter(perfil__aprovado=True).order_by('username')


def _disponibilidade_dispositivos(data):
    linhas = (
        ItemDispositivo.objects
        .filter(agendamento__data=data, agendamento__tipo='DISPOSITIVO')
        .values('agendamento__aula', 'categoria')
        .annotate(total=Sum('quantidade'))
    )
    return {(l['agendamento__aula'], l['categoria']): l['total'] for l in linhas}


def _estoque_por_categoria():
    linhas = (
        Equipamento.objects
        .filter(status='ATIVO', fixo=False)
        .values('categoria')
        .annotate(n=Count('id'))
    )
    return {l['categoria']: l['n'] for l in linhas}


def _datas_futuras_mesmo_dia_semana(data_base):
    """Gera todas as datas do mesmo dia da semana, de data_base até 31/12 do ano."""
    fim_ano = date(data_base.year, 12, 31)
    datas = []
    d = data_base + timedelta(days=7)
    while d <= fim_ano:
        datas.append(d)
        d += timedelta(days=7)
    return datas


def _sobrepor_conflito_sala(data_alvo, aula, sala_id, request_user):
    """
    Remove agendamento de sala conflitante e notifica o professor deslocado.
    Retorna True se havia conflito (e foi removido).
    """
    conflitante = Agendamento.objects.filter(
        data=data_alvo, aula=aula, tipo='SALA', sala_id=sala_id
    ).select_related('professor', 'turma', 'sala').first()

    if not conflitante:
        return False

    sala_nome = conflitante.sala.nome if conflitante.sala else 'sala'
    Notificacao.objects.create(
        destinatario=conflitante.professor,
        mensagem=(
            f'Sua reserva de {data_alvo:%d/%m/%Y} ({aula}ª aula, {sala_nome}) '
            f'foi substituída por um agendamento fixo.'
        ),
    )
    conflitante.delete()
    return True


def estender_agendamentos_fixos():
    """
    Verifica grupos fixos ativos e estende até 31/12 do ano corrente.
    Chamada automaticamente ao acessar a página de agendamentos.
    """
    hoje = date.today()
    fim_ano = date(hoje.year, 12, 31)

    # Encontra grupos fixos cuja última data é anterior ao fim do ano
    grupos = (
        Agendamento.objects
        .filter(fixo=True)
        .exclude(fixo_grupo_id='')
        .values('fixo_grupo_id')
        .annotate(ultima_data=Max('data'))
        .filter(ultima_data__lt=fim_ano)
    )

    for grupo in grupos:
        grupo_id = grupo['fixo_grupo_id']
        ultima_data = grupo['ultima_data']

        # Pega um agendamento modelo do grupo para copiar seus dados
        modelo = (
            Agendamento.objects
            .filter(fixo_grupo_id=grupo_id)
            .select_related('professor', 'turma', 'sala')
            .prefetch_related('itens')
            .order_by('-data')
            .first()
        )
        if not modelo:
            continue

        # Gera datas do dia da semana seguinte ao último até fim do ano
        d = ultima_data + timedelta(days=7)
        while d <= fim_ano:
            if modelo.tipo == 'SALA' and modelo.sala_id:
                _sobrepor_conflito_sala(d, modelo.aula, modelo.sala_id, modelo.professor)
                try:
                    with transaction.atomic():
                        Agendamento.objects.create(
                            data=d, aula=modelo.aula, tipo=modelo.tipo,
                            professor=modelo.professor, turma=modelo.turma,
                            sala=modelo.sala, observacao=modelo.observacao,
                            fixo=True, fixo_grupo_id=grupo_id,
                        )
                except IntegrityError:
                    pass
            elif modelo.tipo == 'DISPOSITIVO':
                try:
                    with transaction.atomic():
                        novo = Agendamento.objects.create(
                            data=d, aula=modelo.aula, tipo='DISPOSITIVO',
                            professor=modelo.professor, turma=modelo.turma,
                            sala=None, observacao=modelo.observacao,
                            fixo=True, fixo_grupo_id=grupo_id,
                        )
                        for item in modelo.itens.all():
                            ItemDispositivo.objects.create(
                                agendamento=novo,
                                categoria=item.categoria,
                                quantidade=item.quantidade,
                            )
                except IntegrityError:
                    pass
            d += timedelta(days=7)


def _processar_agendamento_sala(request, data, ano, mes, dia):
    turma_id = request.POST.get('turma')
    selecionadas = request.POST.getlist('reserva')
    observacao = request.POST.get('observacao', '').strip()
    marcar_fixo = request.POST.get('fixo') == '1' and is_admin_aprovado(request.user)

    if not turma_id:
        messages.warning(request, 'Escolha a turma antes de agendar.')
        return redirect('agendamento_detalhe', ano=ano, mes=mes, dia=dia)
    if not selecionadas:
        messages.warning(request, 'Selecione pelo menos uma sala disponível.')
        return redirect('agendamento_detalhe', ano=ano, mes=mes, dia=dia)

    turma = get_object_or_404(Turma, id=turma_id)
    professor = _resolver_professor(request)
    criados = 0
    conflitos = 0
    grupo_id = str(uuid.uuid4()) if marcar_fixo else ''

    for item in selecionadas:
        try:
            aula_str, sala_str = item.split(':')
            aula = int(aula_str)
            sala_id = int(sala_str)
        except (ValueError, AttributeError):
            continue

        try:
            with transaction.atomic():
                ja_existe = Agendamento.objects.filter(
                    data=data, aula=aula, tipo='SALA', sala_id=sala_id
                ).exists()
                if ja_existe:
                    conflitos += 1
                    continue

                Agendamento.objects.create(
                    data=data, aula=aula, tipo='SALA',
                    professor=professor, turma=turma,
                    sala_id=sala_id, observacao=observacao,
                    fixo=marcar_fixo, fixo_grupo_id=grupo_id,
                )
                criados += 1
        except IntegrityError:
            conflitos += 1

        # Criar agendamentos fixos futuros
        if marcar_fixo:
            datas_futuras = _datas_futuras_mesmo_dia_semana(data)
            for data_futura in datas_futuras:
                _sobrepor_conflito_sala(data_futura, aula, sala_id, request.user)
                try:
                    with transaction.atomic():
                        Agendamento.objects.create(
                            data=data_futura, aula=aula, tipo='SALA',
                            professor=professor, turma=turma,
                            sala_id=sala_id, observacao=observacao,
                            fixo=True, fixo_grupo_id=grupo_id,
                        )
                        criados += 1
                except IntegrityError:
                    pass

    if criados:
        registrar_acao(
            usuario=request.user,
            acao='AGENDOU',
            solicitante_username=professor.username,
            solicitante_email=professor.email,
            tipo_solicitado=professor.perfil.tipo if hasattr(professor, 'perfil') else '',
        )
        if marcar_fixo:
            messages.success(
                request,
                f'{criados} reserva(s) de sala criada(s) (fixo semanal até o final do ano).'
            )
        else:
            messages.success(
                request,
                f'{criados} reserva(s) de sala realizada(s) para {data:%d/%m/%Y}.'
            )
    if conflitos:
        messages.warning(
            request,
            f'{conflitos} sala(s) já tinham sido reservadas nesse meio tempo e foram ignoradas.'
        )
    return redirect('agendamento_detalhe', ano=ano, mes=mes, dia=dia)


def _processar_agendamento_dispositivo(request, data, ano, mes, dia):
    turma_id = request.POST.get('turma')
    observacao = request.POST.get('observacao', '').strip()
    marcar_fixo = request.POST.get('fixo') == '1' and is_admin_aprovado(request.user)

    if not turma_id:
        messages.warning(request, 'Escolha a turma antes de agendar.')
        return redirect('agendamento_detalhe', ano=ano, mes=mes, dia=dia)

    selecao = defaultdict(dict)
    for chave, valor in request.POST.items():
        if not chave.startswith('qtd_'):
            continue
        try:
            _, aula_str, categoria = chave.split('_')
            qtd = int(valor)
        except (ValueError, AttributeError):
            continue
        if qtd > 0:
            selecao[int(aula_str)][categoria] = qtd

    if not selecao:
        messages.warning(request, 'Arraste pelo menos um slider para reservar algum equipamento.')
        return redirect('agendamento_detalhe', ano=ano, mes=mes, dia=dia)

    turma = get_object_or_404(Turma, id=turma_id)
    professor = _resolver_professor(request)
    grupo_id = str(uuid.uuid4()) if marcar_fixo else ''

    aulas_agendadas = 0
    ajustes = 0

    with transaction.atomic():
        reservado = _disponibilidade_dispositivos(data)
        estoque = _estoque_por_categoria()

        for aula, itens in selecao.items():
            itens_validos = []
            for categoria, qtd in itens.items():
                total = estoque.get(categoria, 0)
                disponivel = total - reservado.get((aula, categoria), 0)
                if disponivel <= 0:
                    ajustes += 1
                    continue
                usar = min(qtd, disponivel)
                if usar < qtd:
                    ajustes += 1
                itens_validos.append((categoria, usar))

            if not itens_validos:
                continue

            agendamento = Agendamento.objects.create(
                data=data, aula=aula, tipo='DISPOSITIVO',
                professor=professor, turma=turma,
                sala=None, observacao=observacao,
                fixo=marcar_fixo, fixo_grupo_id=grupo_id,
            )
            for categoria, usar in itens_validos:
                ItemDispositivo.objects.create(
                    agendamento=agendamento, categoria=categoria, quantidade=usar
                )
                reservado[(aula, categoria)] = reservado.get((aula, categoria), 0) + usar
            aulas_agendadas += 1

    # Criar agendamentos fixos futuros para dispositivos
    if marcar_fixo and aulas_agendadas:
        datas_futuras = _datas_futuras_mesmo_dia_semana(data)
        for data_futura in datas_futuras:
            for aula, itens in selecao.items():
                itens_validos = [(cat, qtd) for cat, qtd in itens.items() if qtd > 0]
                if not itens_validos:
                    continue
                try:
                    with transaction.atomic():
                        novo = Agendamento.objects.create(
                            data=data_futura, aula=aula, tipo='DISPOSITIVO',
                            professor=professor, turma=turma,
                            sala=None, observacao=observacao,
                            fixo=True, fixo_grupo_id=grupo_id,
                        )
                        for categoria, usar in itens_validos:
                            ItemDispositivo.objects.create(
                                agendamento=novo, categoria=categoria, quantidade=usar
                            )
                        aulas_agendadas += 1
                except IntegrityError:
                    pass

    if aulas_agendadas:
        registrar_acao(
            usuario=request.user,
            acao='AGENDOU',
            solicitante_username=professor.username,
            solicitante_email=professor.email,
            tipo_solicitado=professor.perfil.tipo if hasattr(professor, 'perfil') else '',
        )
        if marcar_fixo:
            messages.success(
                request,
                f'Equipamentos reservados em {aulas_agendadas} aula(s) (fixo semanal até o final do ano).'
            )
        else:
            messages.success(
                request,
                f'Equipamentos reservados em {aulas_agendadas} aula(s) no dia {data:%d/%m/%Y}.'
            )
    if ajustes:
        messages.warning(
            request,
            f'{ajustes} item(ns) reduzido(s) ou ignorado(s) por falta de estoque disponível.'
        )
    if not aulas_agendadas:
        messages.warning(request, 'Nenhum equipamento pôde ser reservado (estoque esgotado).')
    return redirect('agendamento_detalhe', ano=ano, mes=mes, dia=dia)


@login_required
def agendamentos(request):
    """Exibe o calendário mensal com visão de agendamentos e navegação reativa."""
    if not is_usuario_aprovado(request.user):
        return redirect('home')

    # Extensão dinâmica: estende agendamentos fixos até 31/12 do ano corrente
    estender_agendamentos_fixos()

    hoje = date.today()

    try:
        ano = int(request.GET.get('ano', hoje.year))
        mes = int(request.GET.get('mes', hoje.month))
        dia = int(request.GET.get('dia', hoje.day))
        data_atual = date(ano, mes, dia)
    except (TypeError, ValueError):
        ano, mes, dia = hoje.year, hoje.month, hoje.day
        data_atual = hoje

    if not (1 <= mes <= 12):
        mes = hoje.month

    dia_ant = data_atual - timedelta(days=1)
    dia_prox = data_atual + timedelta(days=1)

    mes_ant = data_atual.month - 1
    ano_mes_ant = data_atual.year
    if mes_ant < 1:
        mes_ant = 12
        ano_mes_ant -= 1

    mes_prox = data_atual.month + 1
    ano_mes_prox = data_atual.year
    if mes_prox > 12:
        mes_prox = 1
        ano_mes_prox += 1

    ano_ant = data_atual.year - 1
    ano_prox = data_atual.year + 1

    cal = calendar.Calendar(firstweekday=6)
    semanas = []
    for semana in cal.monthdayscalendar(ano, mes):
        linha = []
        for d in semana:
            if d == 0:
                linha.append(None)
            else:
                dia_data = date(ano, mes, d)
                linha.append({
                    'numero': d,
                    'hoje': (dia_data == hoje),
                    'passado': (dia_data < hoje),
                })
        semanas.append(linha)

    dias_para_domingo = data_atual.isoweekday() % 7
    domingo = data_atual - timedelta(days=dias_para_domingo)
    sabado = domingo + timedelta(days=6)

    reservas_qs = (
        Agendamento.objects.filter(data__range=(domingo, sabado))
        .select_related('sala', 'turma', 'professor')
        .prefetch_related('itens')
    )

    dias_cabecalho = []
    for i in range(7):
        d = domingo + timedelta(days=i)
        dias_cabecalho.append({
            'data': d,
            'numero': d.day,
            'mes_nome': MESES_PT[d.month - 1][:3],
            'nome_curto': DIAS_SEMANA_LONGO[d.weekday()][:3],
            'hoje': d == hoje
        })

    grade_semanal = []
    for aula in range(1, 10):
        linha = []
        for i in range(7):
            d = domingo + timedelta(days=i)
            reservas_slot = [r for r in reservas_qs if r.data == d and r.aula == aula]
            tem_reserva_usuario = any(r.professor == request.user for r in reservas_slot)
            linha.append({
                'data': d,
                'reservas': reservas_slot,
                'tem_reserva_usuario': tem_reserva_usuario
            })
        grade_semanal.append({'aula': aula, 'dias': linha})

    context = {
        'title': 'Agendamentos',
        'data_atual': data_atual,
        'ano': ano,
        'mes': mes,
        'dia': dia,
        'mes_nome': MESES_PT[mes - 1],
        'dias_semana': DIAS_SEMANA_PT,
        'semanas': semanas,
        'dia_ant': dia_ant, 'dia_prox': dia_prox,
        'mes_ant': mes_ant, 'ano_mes_ant': ano_mes_ant,
        'mes_prox': mes_prox, 'ano_mes_prox': ano_mes_prox,
        'ano_ant': ano_ant, 'ano_prox': ano_prox,
        'hoje_ano': hoje.year,
        'hoje_mes': hoje.month,
        'hoje_dia': hoje.day,
        'lista_meses': list(enumerate(MESES_PT, start=1)),
        'lista_anos': range(hoje.year - 2, hoje.year + 4),
        'is_admin': is_admin_aprovado(request.user),
        'grade_semanal': grade_semanal,
        'dias_cabecalho': dias_cabecalho,
        'domingo': domingo,
        'sabado': sabado,
    }
    return render(request, 'app/agendamentos.html', context)

@login_required
def exportar_excel_mes(request):
    """Gera e baixa um arquivo Excel com os agendamentos semanais do mês escolhido."""
    if not is_usuario_aprovado(request.user):
        return redirect('home')

    hoje = date.today()
    try:
        ano = int(request.GET.get('ano', hoje.year))
        mes = int(request.GET.get('mes', hoje.month))
    except ValueError:
        ano, mes = hoje.year, hoje.month

    if not (1 <= mes <= 12):
        mes = hoje.month

    nome_mes = MESES_PT[mes - 1]

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Agendamentos_{nome_mes}_{ano}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = f"{nome_mes.capitalize()} {ano}"

    fonte_titulo = Font(bold=True, size=14, color="FFFFFF")
    fonte_cabecalho = Font(bold=True, size=11, color="FFFFFF")
    preenchimento_titulo = PatternFill(start_color="1A4A8A", end_color="1A4A8A", fill_type="solid")
    preenchimento_cabecalho = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
    preenchimento_aula = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    alinhamento_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alinhamento_topo = Alignment(horizontal="left", vertical="top", wrap_text=True)
    borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Título Principal
    ws.append([f"Agendamentos - {nome_mes.capitalize()} de {ano}"])
    ws.merge_cells('A1:H1')
    celula_titulo = ws.cell(row=1, column=1)
    celula_titulo.font = fonte_titulo
    celula_titulo.fill = preenchimento_titulo
    celula_titulo.alignment = alinhamento_centro
    ws.append([]) # Linha em branco

    linha_atual = 3
    cal = calendar.Calendar(firstweekday=6) # Domingo = 6

    for semana in cal.monthdatescalendar(ano, mes):
        if not any(d.month == mes for d in semana):
            continue

        domingo = semana[0]
        sabado = semana[-1]
        
        # Subtítulo da Semana
        texto_semana = f"Semana de {domingo.strftime('%d/%m')} a {sabado.strftime('%d/%m')}"
        ws.append([texto_semana])
        ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=8)
        cel_semana = ws.cell(row=linha_atual, column=1)
        cel_semana.font = Font(bold=True, size=12)
        linha_atual += 1

        reservas_semana = (
            Agendamento.objects.filter(data__range=(domingo, sabado))
            .select_related('sala', 'turma', 'professor')
        )

        cabecalho = ['Aula'] + [f"{DIAS_SEMANA_PT[d.isoweekday()%7]} {d.day}" for d in semana]
        ws.append(cabecalho)
        
        # Estilizar cabeçalho
        for col_idx in range(1, 9):
            cel = ws.cell(row=linha_atual, column=col_idx)
            cel.font = fonte_cabecalho
            cel.fill = preenchimento_cabecalho
            cel.alignment = alinhamento_centro
            cel.border = borda_fina
        
        linha_atual += 1

        for aula in range(1, 10):
            linha_dados = [f"{aula}ª"]
            for d in semana:
                reservas_slot = [r for r in reservas_semana if r.data == d and r.aula == aula]
                if not reservas_slot:
                    linha_dados.append("")
                else:
                    textos = []
                    for r in reservas_slot:
                        local = r.sala.nome if r.tipo == 'SALA' else 'Equip'
                        textos.append(f"{local} - {r.turma.nome} ({r.professor.username})")
                    linha_dados.append("\n".join(textos))
            
            ws.append(linha_dados)
            
            # Estilizar dados da aula
            for col_idx in range(1, 9):
                cel = ws.cell(row=linha_atual, column=col_idx)
                cel.border = borda_fina
                if col_idx == 1:
                    cel.font = Font(bold=True)
                    cel.fill = preenchimento_aula
                    cel.alignment = alinhamento_centro
                else:
                    cel.alignment = alinhamento_topo
            
            # Ajustar altura da linha baseado no conteúdo
            max_quebras = max([item.count('\n') for item in linha_dados]) if any(linha_dados) else 0
            ws.row_dimensions[linha_atual].height = 15 + (max_quebras * 15)
            
            linha_atual += 1
            
        ws.append([]) # Linha em branco entre semanas
        linha_atual += 1

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 10
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letter].width = 25

    wb.save(response)
    return response


@login_required
def agendamento_detalhe(request, ano, mes, dia):
    """Exibe detalhes do dia selecionado e processa novas reservas."""
    if not is_usuario_aprovado(request.user):
        return redirect('home')

    try:
        data = date(ano, mes, dia)
    except ValueError:
        messages.error(request, 'Data inválida.')
        return redirect('agendamentos')

    is_admin = is_admin_aprovado(request.user)
    if data < date.today() and not is_admin:
        messages.warning(
            request,
            f'{data:%d/%m/%Y} é uma data passada — não é possível agendar.'
        )
        return redirect('agendamentos')

    if request.method == 'POST':
        if request.POST.get('tipo') == 'sala':
            return _processar_agendamento_sala(request, data, ano, mes, dia)
        if request.POST.get('tipo') == 'dispositivo':
            return _processar_agendamento_dispositivo(request, data, ano, mes, dia)
        return redirect('agendamento_detalhe', ano=ano, mes=mes, dia=dia)

    salas_ativas = list(Sala.objects.filter(ativo=True))

    reservas = (
        Agendamento.objects
        .filter(data=data, tipo='SALA')
        .select_related('sala', 'turma', 'professor')
    )
    ocupacao = {(r.aula, r.sala_id): r for r in reservas}

    aulas = []
    for numero, (nome, horario) in enumerate(AULAS_HORARIOS, start=1):
        linha_salas = []
        livres = 0
        for sala in salas_ativas:
            reserva = ocupacao.get((numero, sala.id))
            if reserva:
                prof = reserva.professor.get_full_name() or reserva.professor.username
                linha_salas.append({
                    'sala': sala, 'ocupado': True,
                    'turma': reserva.turma.nome, 'professor': prof,
                    'ag_id': reserva.id,
                })
            else:
                linha_salas.append({'sala': sala, 'ocupado': False})
                livres += 1
        aulas.append({
            'numero': numero, 'nome': nome, 'horario': horario,
            'salas': linha_salas, 'livres': livres, 'total': len(salas_ativas),
        })

    estoque_categoria = _estoque_por_categoria()
    reservado_disp = _disponibilidade_dispositivos(data)

    ags_disp = (
        Agendamento.objects
        .filter(data=data, tipo='DISPOSITIVO')
        .select_related('turma', 'professor')
        .prefetch_related('itens')
        .order_by('aula')
    )
    cards_por_aula = defaultdict(list)
    for ag in ags_disp:
        cards_por_aula[ag.aula].append(ag)

    aulas_disp = []
    for numero, (nome, horario) in enumerate(AULAS_HORARIOS, start=1):
        grupos = []
        for cat_valor, cat_label in Equipamento.CATEGORIA_CHOICES:
            total = estoque_categoria.get(cat_valor, 0)
            if total == 0:
                continue
            restante = max(total - reservado_disp.get((numero, cat_valor), 0), 0)
            ratio = restante / total if total else 0
            if restante == 0:
                status = 'vermelho'
            elif ratio <= 0.5:
                status = 'laranja'
            else:
                status = 'verde'
            grupos.append({
                'categoria': cat_valor, 'label': cat_label,
                'restante': restante, 'total': total, 'status': status,
            })
        aulas_disp.append({
            'numero': numero, 'nome': nome, 'horario': horario,
            'grupos': grupos,
            'cards': cards_por_aula.get(numero, []),
        })

    context = {
        'title': f'Agendamento {data:%d/%m/%Y}',
        'data': data,
        'ano': ano, 'mes': mes, 'dia': dia,
        'dia_semana': DIAS_SEMANA_LONGO[data.weekday()],
        'mes_nome': MESES_PT[mes - 1],
        'aulas': aulas,
        'aulas_disp': aulas_disp,
        'turmas': Turma.objects.all(),
        'tem_salas': bool(salas_ativas),
        'tem_equipamentos': bool(estoque_categoria),
        'is_admin': is_admin,
        'professores': _professores_aprovados() if is_admin else None,
    }
    return render(request, 'app/agendamento_detalhe.html', context)


@login_required
def cancelar_reserva(request, agendamento_id):
    """Cancela uma reserva e registra no histórico de auditoria."""
    if not is_usuario_aprovado(request.user):
        return redirect('home')

    if request.method == 'POST':
        try:
            ag = get_object_or_404(Agendamento, id=agendamento_id)
            is_admin = is_admin_aprovado(request.user)

            if ag.professor != request.user and not is_admin:
                msg = 'Você só pode cancelar as suas próprias reservas.'
                if is_ajax(request):
                    return JsonResponse({'ok': False, 'message': msg})
                messages.error(request, msg)
                return redirect('agendamentos')

            prof = ag.professor
            cancelar_tipo = request.POST.get('cancelar_tipo', 'hoje')

            if cancelar_tipo == 'todos' and ag.fixo and ag.fixo_grupo_id:
                # Cancelar todos os futuros do grupo fixo
                removidos = Agendamento.objects.filter(
                    fixo_grupo_id=ag.fixo_grupo_id,
                    data__gte=date.today(),
                ).delete()[0]
                msg = f'{removidos} reserva(s) fixa(s) cancelada(s).'
            else:
                ag.delete()
                msg = 'Reserva cancelada com sucesso.'

            registrar_acao(
                usuario=request.user,
                acao='CANCELOU_AGENDAMENTO',
                solicitante_username=prof.username,
                solicitante_email=prof.email,
                tipo_solicitado=prof.perfil.tipo if hasattr(prof, 'perfil') else '',
            )

            # Enviar notificação ao professor caso o admin cancele
            if request.user != prof:
                Notificacao.objects.create(
                    destinatario=prof,
                    mensagem=f'Sua reserva do dia {ag.data.strftime("%d/%m/%Y")} ({ag.aula}ª aula) foi cancelada pelo administrador {request.user.get_full_name() or request.user.username}.'
                )

            if is_ajax(request):
                restantes = Agendamento.objects.filter(
                    professor=request.user, data__gte=date.today()
                ).count()
                return JsonResponse({
                    'ok': True, 'acao': 'recarregar_pagina' if cancelar_tipo == 'todos' else 'remover_linha', 'message': msg,
                    'restantes': restantes,
                    'vazio_html': (
                        '<tr><td colspan="6" class="text-center text-muted" style="padding: 32px;">'
                        '<span>📭</span>'
                        '<p style="margin-top: 8px; font-style: italic;">Você não tem reservas futuras. '
                        'Clique em um dia para agendar.</p></td></tr>'
                    ),
                })
            messages.warning(request, msg)
        except Exception as e:
            import traceback
            error_msg = f'Erro interno: {str(e)}\n\n{traceback.format_exc()}'
            if is_ajax(request):
                return JsonResponse({'ok': False, 'message': error_msg})
            messages.error(request, error_msg)

    return redirect('agendamentos')


def _salas_para_edicao(ag):
    ocupadas = set(
        Agendamento.objects
        .filter(data=ag.data, aula=ag.aula, tipo='SALA')
        .exclude(id=ag.id)
        .values_list('sala_id', flat=True)
    )
    opcoes = []
    for s in Sala.objects.filter(ativo=True):
        if s.id not in ocupadas or s.id == ag.sala_id:
            opcoes.append({'sala': s, 'atual': s.id == ag.sala_id})
    return opcoes


def _categorias_para_edicao(ag):
    estoque = _estoque_por_categoria()
    reservado_total = _disponibilidade_dispositivos(ag.data)
    atuais = {it.categoria: it.quantidade for it in ag.itens.all()}

    opcoes = []
    for cat_valor, cat_label in Equipamento.CATEGORIA_CHOICES:
        total = estoque.get(cat_valor, 0)
        atual = atuais.get(cat_valor, 0)
        if total == 0 and atual == 0:
            continue
        reservado_outros = reservado_total.get((ag.aula, cat_valor), 0) - atual
        maximo = max(total - reservado_outros, 0)
        opcoes.append({
            'categoria': cat_valor, 'label': cat_label,
            'atual': atual, 'maximo': maximo,
        })
    return opcoes


def _aplicar_edicao_dispositivo(request, ag):
    estoque = _estoque_por_categoria()
    reservado_total = _disponibilidade_dispositivos(ag.data)
    atuais = {it.categoria: it for it in ag.itens.all()}

    for cat_valor, _label in Equipamento.CATEGORIA_CHOICES:
        campo = request.POST.get(f'qtd_cat_{cat_valor}')
        if campo is None:
            continue
        try:
            novo = max(int(campo), 0)
        except (TypeError, ValueError):
            continue

        atual_qtd = atuais[cat_valor].quantidade if cat_valor in atuais else 0
        reservado_outros = reservado_total.get((ag.aula, cat_valor), 0) - atual_qtd
        maximo = max(estoque.get(cat_valor, 0) - reservado_outros, 0)
        novo = min(novo, maximo)

        if novo > 0:
            if cat_valor in atuais:
                item = atuais[cat_valor]
                item.quantidade = novo
                item.save()
            else:
                ItemDispositivo.objects.create(
                    agendamento=ag, categoria=cat_valor, quantidade=novo
                )
        elif cat_valor in atuais:
            atuais[cat_valor].delete()


@login_required
def relacao_agendamento(request, agendamento_id):
    """Gerencia a atribuição de equipamentos aos alunos da turma reservada."""
    if not is_usuario_aprovado(request.user):
        return redirect('home')

    ag = get_object_or_404(
        Agendamento.objects
        .select_related('sala', 'turma', 'professor')
        .prefetch_related('itens'),
        id=agendamento_id
    )

    is_admin = is_admin_aprovado(request.user)
    pode_editar = is_admin or ag.professor == request.user
    alunos = list(ag.turma.alunos.all())

    if request.method == 'POST':
        if not pode_editar:
            messages.error(request, 'Você não pode editar esta reserva.')
            return redirect('relacao_agendamento', agendamento_id=ag.id)

        acao = request.POST.get('acao', 'relacao')

        if acao == 'editar':
            ag.observacao = request.POST.get('observacao', '').strip()

            turma = Turma.objects.filter(id=request.POST.get('turma')).first()
            if turma and turma != ag.turma:
                ag.relacoes.all().delete()
                ag.turma = turma

            if is_admin:
                prof = User.objects.filter(id=request.POST.get('professor')).first()
                if prof:
                    ag.professor = prof

            if ag.tipo == 'SALA':
                nova = request.POST.get('sala', '')
                if nova.isdigit() and int(nova) != ag.sala_id:
                    nova_id = int(nova)
                    ocupada = (
                        Agendamento.objects
                        .filter(data=ag.data, aula=ag.aula, tipo='SALA', sala_id=nova_id)
                        .exclude(id=ag.id).exists()
                    )
                    if ocupada:
                        messages.warning(request, 'A sala escolhida já está ocupada nessa aula; mantida a anterior.')
                    else:
                        ag.sala_id = nova_id

            ag.save()

            if ag.tipo == 'DISPOSITIVO':
                _aplicar_edicao_dispositivo(request, ag)
                if not ag.itens.exists():
                    ag.delete()
                    messages.warning(request, 'A reserva ficou sem equipamentos e foi removida.')
                    return redirect('agendamentos')

            registrar_acao(
                usuario=request.user,
                acao='ALTEROU_AGENDAMENTO',
                solicitante_username=ag.professor.username,
                solicitante_email=ag.professor.email,
                tipo_solicitado=ag.professor.perfil.tipo if hasattr(ag.professor, 'perfil') else '',
            )

            messages.success(request, 'Reserva atualizada com sucesso!')
            return redirect('relacao_agendamento', agendamento_id=ag.id)

        for aluno in alunos:
            valor = request.POST.get(f'equip_{aluno.id}', '').strip()
            RelacaoAlunoEquipamento.objects.update_or_create(
                agendamento=ag, aluno=aluno,
                defaults={'equipamento': valor},
            )
        messages.success(request, 'Relação de alunos e equipamentos salva com sucesso!')
        return redirect('relacao_agendamento', agendamento_id=ag.id)

    salvos = {r.aluno_id: r.equipamento for r in ag.relacoes.all()}
    linhas = [{'aluno': a, 'equipamento': salvos.get(a.id, '')} for a in alunos]

    return render(request, 'app/relacao_agendamento.html', {
        'title': 'Relação Alunos x Equipamentos',
        'ag': ag,
        'linhas': linhas,
        'pode_editar': pode_editar,
        'is_admin': is_admin,
        'turmas': Turma.objects.all(),
        'professores': _professores_aprovados() if is_admin else None,
        'salas_edicao': _salas_para_edicao(ag) if pode_editar and ag.tipo == 'SALA' else None,
        'categorias_edicao': _categorias_para_edicao(ag) if pode_editar and ag.tipo == 'DISPOSITIVO' else None,
    })
